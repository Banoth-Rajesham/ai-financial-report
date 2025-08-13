# ==============================================================================
# FILE: agents/agent_5_reporter.py (FINAL, WITH BORDERS, COLORS & FORMULAS)
# ==============================================================================
import pandas as pd
import io
import traceback
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from config import MASTER_TEMPLATE, NOTES_STRUCTURE_AND_MAPPING

def apply_main_sheet_styling(ws, template, company_name, note_to_row_map):
    """Applies the final, professional styling with borders, colors, and formulas."""

    # --- STYLE DEFINITIONS ---
    main_header_green = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
    sub_header_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    column_header_grey = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    bold_font = Font(bold=True)
    
    currency_format = '#,##0.00'
    
    thin_border_side = Side(style='thin', color="BFBFBF")
    full_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # --- APPLY STYLES ---
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 65
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20

    ws.insert_rows(1, 3)
    ws.merge_cells('A1:E1')
    ws['A1'] = company_name
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:E2')
    ws['A2'] = ws.title
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = Alignment(horizontal='center')

    for cell in ws[4]:
        cell.fill = column_header_grey
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = full_border

    for i, row_template in enumerate(template, start=1):
        row_num = i + 4
        row_type = row_template[3]
        notes_to_sum = row_template[2]

        if row_type != 'spacer':
            for col in range(1, 6):
                ws.cell(row=row_num, column=col).border = full_border

            if row_type == 'header':
                ws[f'B{row_num}'].font = bold_font
                for col in ['A', 'B', 'C', 'D', 'E']: ws[f'{col}{row_num}'].fill = main_header_green

            elif row_type == 'sub_header':
                ws[f'B{row_num}'].font = bold_font
                for col in ['A', 'B', 'C', 'D', 'E']: ws[f'{col}{row_num}'].fill = sub_header_blue

            elif row_type == 'total':
                ws[f'B{row_num}'].font = bold_font
                for col in ['A', 'B', 'C', 'D', 'E']: ws[f'{col}{row_num}'].fill = main_header_green
                
                # --- FORMULA GENERATION ---
                if isinstance(notes_to_sum, list):
                    sum_cells_d = [f'D{note_to_row_map[n]}' for n in notes_to_sum if n in note_to_row_map]
                    if sum_cells_d: ws[f'D{row_num}'] = f"=SUM({','.join(sum_cells_d)})"
                    
                    sum_cells_e = [f'E{note_to_row_map[n]}' for n in notes_to_sum if n in note_to_row_map]
                    if sum_cells_e: ws[f'E{row_num}'] = f"=SUM({','.join(sum_cells_e)})"

            for col_letter in ['D', 'E']:
                cell = ws[f'{col_letter}{row_num}']
                if cell.value is not None and isinstance(cell.value, (int, float, str)) and "=" not in str(cell.value):
                     if isinstance(cell.value, (int, float)):
                        cell.number_format = currency_format


def apply_note_sheet_styling(ws, note_title):
    """Applies professional styling with borders to a Note sheet."""
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    total_font = Font(bold=True)
    currency_format = '#,##0.00'
    
    thin_border_side = Side(style='thin', color="BFBFBF")
    full_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    ws.column_dimensions['A'].width = 65
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    ws.merge_cells('A1:C1')
    ws['A1'] = note_title
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = full_border

    for row in ws.iter_rows(min_row=3, max_col=3):
        is_total_row = (row[0].value == 'Total')
        for cell in row:
            cell.border = full_border
            if cell.column > 1 and isinstance(cell.value, (int, float)):
                cell.number_format = currency_format
            if is_total_row:
                cell.font = total_font
                cell.border = Border(top=Side(style='double', color="BFBFBF"), left=thin_border_side, right=thin_border_side, bottom=thin_border_side)


def report_finalizer_agent(aggregated_data, company_name):
    """AGENT 5: Constructs the final report with styling, borders, and formulas."""
    print("\n--- Agent 5 (Report Finalizer): Building styled report... ---")
    try:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:

            for sheet_name, template in [("Balance Sheet", MASTER_TEMPLATE["Balance Sheet"]),
                                         ("Profit and Loss", MASTER_TEMPLATE["Profit and Loss"])]:
                
                sheet_data = []
                note_to_row_map = {}
                pbt_cy, pbt_py, total_revenue_cy, total_revenue_py, total_expenses_cy, total_expenses_py, total_tax_cy, total_tax_py = [0]*8

                for index, row_template in enumerate(template):
                    excel_row_num = index + 5
                    col_a, particulars, note, row_type = row_template
                    row = { ' ': col_a, 'Particulars': particulars, 'Note': "" if not isinstance(note, str) else note,
                            'As at March 31, 2025': None, 'As at March 31, 2024': None }

                    if row_type in ["item", "item_no_alpha"]:
                        note_str = str(note)
                        note_to_row_map[note_str] = excel_row_num
                        row['As at March 31, 2025'] = aggregated_data.get(note_str, {}).get('total', {}).get('CY', 0)
                        row['As at March 31, 2024'] = aggregated_data.get(note_str, {}).get('total', {}).get('PY', 0)
                    elif row_type == "total" and isinstance(note, list):
                        cy_val = sum(aggregated_data.get(str(n), {}).get('total', {}).get('CY', 0) for n in note)
                        py_val = sum(aggregated_data.get(str(n), {}).get('total', {}).get('PY', 0) for n in note)
                        row['As at March 31, 2025'], row['As at March 31, 2024'] = cy_val, py_val
                        if "Revenue" in particulars: total_revenue_cy, total_revenue_py = cy_val, py_val
                        if "Expenses" in particulars: total_expenses_cy, total_expenses_py = cy_val, py_val
                        if "Tax Expense" in particulars: total_tax_cy, total_tax_py = cy_val, py_val
                    elif note == "PBT":
                        pbt_cy, pbt_py = (total_revenue_cy - total_expenses_cy), (total_revenue_py - total_expenses_py)
                        row['As at March 31, 2025'], row['As at March 31, 2024'] = pbt_cy, pbt_py
                    elif note == "PAT":
                        row['As at March 31, 2025'], row['As at March 31, 2024'] = (pbt_cy - total_tax_cy), (pbt_py - total_tax_py)
                    
                    sheet_data.append(row)

                df = pd.DataFrame(sheet_data)
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=True, startrow=3)
                ws = writer.sheets[sheet_name]
                apply_main_sheet_styling(ws, template, company_name, note_to_row_map)

            for note_num_str in sorted(NOTES_STRUCTURE_AND_MAPPING.keys(), key=int):
                note_info = NOTES_STRUCTURE_AND_MAPPING[note_num_str]
                note_data = aggregated_data.get(note_num_str)
                if note_data and note_data.get('sub_items'):
                    sheet_name = f'Note {note_num_str}'
                    note_title = f'Note {note_num_str}: {note_info["title"]}'
                    note_df_data = []

                    def process_sub_items(items, level=0):
                        for key, value in items.items():
                            indent = '  ' * level
                            if isinstance(value, dict) and 'CY' in value:
                                note_df_data.append({'Particulars': indent + key, 'As at March 31, 2025': value.get('CY', 0), 'As at March 31, 2024': value.get('PY', 0)})
                            elif isinstance(value, dict):
                                note_df_data.append({'Particulars': indent + key, 'As at March 31, 2025': None, 'As at March 31, 2024': None})
                                process_sub_items(value, level + 1)
                    
                    process_sub_items(note_data['sub_items'])
                    note_df = pd.DataFrame(note_df_data)
                    total_row = pd.DataFrame([{'Particulars': 'Total', 'As at March 31, 2025': note_data['total'].get('CY', 0), 'As at March 31, 2024': note_data['total'].get('PY', 0)}])
                    note_df = pd.concat([note_df, total_row], ignore_index=True)
                    
                    note_df.to_excel(writer, sheet_name=sheet_name, index=False, header=True, startrow=1)
                    ws_note = writer.sheets[sheet_name]
                    apply_note_sheet_styling(ws_note, note_title)

        print("✅ Report Finalizer SUCCESS: Report generated with styling.")
        return output.getvalue()
    except Exception as e:
        print(f"❌ Report Finalizer FAILED: {e}")
        traceback.print_exc()
        return None
