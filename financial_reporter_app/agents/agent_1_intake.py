# ==============================================================================
# PASTE THIS ENTIRE, CORRECTED BLOCK INTO: agent_1_intake.py
# ==============================================================================
import pandas as pd
import streamlit as st
import openpyxl # Use a more powerful library for scanning

def intelligent_data_intake_agent(uploaded_file):
    """
    AGENT 1: Ingests the uploaded Excel file with a highly robust method that
    scans all sheets to find the header row and standardizes columns.
    This version is immune to merged cells and data being on any sheet.
    """
    print("\n--- Agent 1 (Data Intake): Ingesting file... ---")
    if uploaded_file is None:
        return None

    try:
        # --- THIS IS THE PERMANENT FIX ---
        workbook = openpyxl.load_workbook(uploaded_file, read_only=True)
        
        found_header = False
        target_sheet_name = None
        header_row_idx = -1
        particulars_col_idx = -1

        # 1. Iterate through ALL sheets in the workbook to find the right one
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # 2. Scan the first 30 rows of the current sheet
            for r_idx in range(1, 31):
                for c_idx in range(1, 31): # Scan first 30 columns as well
                    cell = sheet.cell(row=r_idx, column=c_idx)
                    if cell.value and isinstance(cell.value, str) and 'particulars' in cell.value.lower():
                        # Found it! Store the details.
                        header_row_idx = r_idx - 1 # 0-indexed for pandas
                        particulars_col_idx = c_idx - 1 # 0-indexed for pandas
                        target_sheet_name = sheet_name
                        found_header = True
                        break
                if found_header:
                    break
            if found_header:
                break

        if not found_header:
            st.error("Data Intake Error: Scanned all sheets, but could not find a cell containing the word 'Particulars'. Please check the file.")
            return None

        # 3. Read the file again, but this time using the correct sheet and header row
        uploaded_file.seek(0)
        df = pd.read_excel(uploaded_file, sheet_name=target_sheet_name, header=header_row_idx)

        # 4. Get the actual column name from the found index and standardize
        original_cols = df.columns
        particulars_col_name = original_cols[particulars_col_idx]
        
        if len(original_cols) < particulars_col_idx + 4:
             st.error(f"Data Intake Error: Found 'Particulars' in column {particulars_col_idx+1}, but the file is missing subsequent Note/Amount columns.")
             return None
             
        note_col_name = original_cols[particulars_col_idx + 1]
        cy_col_name = original_cols[particulars_col_idx + 2]
        py_col_name = original_cols[particulars_col_idx + 3]

        df_clean = df[[particulars_col_name, note_col_name, cy_col_name, py_col_name]].copy()
        df_clean.columns = ['Particulars', 'Note', 'Amount_CY', 'Amount_PY']

        df_clean = df_clean.dropna(subset=['Particulars'])
        df_clean = df_clean[df_clean['Particulars'].astype(str).str.lower() != 'particulars']
        
        print(f"✅ Data Intake SUCCESS: Found header on sheet '{target_sheet_name}' at row {header_row_idx+1}. Columns standardized.")
        return df_clean

    except Exception as e:
        st.error(f"Data Intake FAILED: An unexpected error occurred. Please ensure the Excel file is not corrupted. Error: {e}")
        import traceback
        traceback.print_exc()
        return None
